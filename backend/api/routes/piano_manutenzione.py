"""
Piano di Manutenzione — routes.

Gerarchia: Asset → Piano → Task → Ticket

Endpoint principali:
  GET    /piani-manutenzione                  lista piani (con conteggi task/ticket)
  POST   /piani-manutenzione                  crea piano (asset_id obbligatorio)
  GET    /piani-manutenzione/{id}             dettaglio piano
  PUT    /piani-manutenzione/{id}             aggiorna piano
  DELETE /piani-manutenzione/{id}             elimina piano (scollegando task e ticket)

  GET    /piani-manutenzione/{id}/tasks       lista task del piano
  POST   /piani-manutenzione/{id}/tasks       crea task nel piano
  PUT    /piani-manutenzione/{id}/tasks/{tid} aggiorna task
  DELETE /piani-manutenzione/{id}/tasks/{tid} elimina task dal piano

  POST   /piani-manutenzione/{id}/tasks/{tid}/genera-ticket  genera ticket da task
  GET    /piani-manutenzione/{id}/tasks/{tid}/tickets        storico ticket del task

  POST   /piani-manutenzione/{id}/import-pdf   import PDF → task
  POST   /piani-manutenzione/{id}/import-excel import Excel → task
"""
import io
import csv
import math
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func

from backend.core.dependencies import get_db
from backend.core.security import get_current_tenant_id, check_tenant_ownership
from backend.core.logger_db import db_info, db_error
from backend.db.modelli import PianoManutenzione, Ticket, AttivitaManutenzione, Asset, Manuale, piano_asset_association
from backend.services.condition_maintenance_service import (
    METRIC_RUNNING_HOURS,
    latest_running_hours_by_asset,
    normalize_trigger_mode,
    task_due_summary,
)
from backend.schemas.piano_manutenzione import (
    PianoManutenzioneCreate, PianoManutenzioneUpdate, PianoManutenzioneResponse,
    PianoManutenzioneListItem, TaskCreate, TaskUpdate, TaskResponse,
    TicketPianoAssign,
)

router = APIRouter()

_PRIO_MAP = {"high": "Alta", "medium": "Media", "low": "Bassa",
             "alta": "Alta", "media": "Media", "bassa": "Bassa"}
MAX_IMPORT_BYTES = 25 * 1024 * 1024  # 25 MB


def _freq_days(freq: dict | None) -> int | None:
    if not freq:
        return None
    value = freq.get("value")
    if value is None:
        return None
    unit = (freq.get("unit") or "days").lower()
    multipliers = {"days": 1, "weeks": 7, "months": 30, "years": 365}
    return int(value * multipliers.get(unit, 1))


def _task_to_dict(
    t: AttivitaManutenzione,
    open_ticket_count: int = 0,
    asset_nome: str | None = None,
    latest_reading=None,
) -> dict:
    """Serializza un task con informazioni aggiuntive."""
    next_due = None
    if t.last_generated_at and t.frequenza_giorni:
        next_due = t.last_generated_at + timedelta(days=t.frequenza_giorni)
    elif t.prossima_scadenza:
        next_due = t.prossima_scadenza
    elif t.next_due_at:
        next_due = t.next_due_at

    due = task_due_summary(t, latest_reading)
    condition = due["condition"]

    return {
        "id": t.id,
        "piano_id": getattr(t, "piano_id", None),
        "asset_id": t.asset_id,
        "asset_nome": asset_nome or "",
        "nome": getattr(t, "nome", None) or "",
        "descrizione": t.descrizione or "",
        "frequenza_giorni": t.frequenza_giorni,
        "durata_ore": t.durata_ore,
        "priorita": t.priorita or "Media",
        "codice": getattr(t, "codice", None),
        "is_repeatable": getattr(t, "is_repeatable", True),
        "generation_mode": getattr(t, "generation_mode", "manual") or "manual",
        "generate_days_before_due": getattr(t, "generate_days_before_due", 7),
        "task_stato": getattr(t, "task_stato", "active") or "active",
        "source_type": getattr(t, "source_type", None),
        "last_generated_at": t.last_generated_at.isoformat() if t.last_generated_at else None,
        "next_due_at": next_due.isoformat() if next_due else None,
        "open_ticket_count": open_ticket_count,
        "trigger_mode": normalize_trigger_mode(getattr(t, "trigger_mode", None)),
        "condition_metric": getattr(t, "condition_metric", None),
        "condition_threshold_hours": getattr(t, "condition_threshold_hours", None),
        "condition_last_done_hours": getattr(t, "condition_last_done_hours", None),
        "current_running_hours": condition.get("current_hours"),
        "condition_due_at_hours": condition.get("due_at_hours"),
        "condition_remaining_hours": condition.get("remaining_hours"),
        "condition_is_due": condition.get("is_due", False),
    }


# ── PIANI ─────────────────────────────────────────────────────────────────────

@router.get("/piani-manutenzione")
def list_piani_manutenzione(
    asset_id: Optional[int] = Query(None),
    stato: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Lista piani con conteggi task e ticket aperti."""
    query = db.query(PianoManutenzione).filter(PianoManutenzione.tenant_id == tenant_id)
    if asset_id is not None:
        query = query.filter(PianoManutenzione.asset_id == asset_id)
    if stato:
        query = query.filter(PianoManutenzione.stato == stato)

    total = query.count()
    piani = query.order_by(PianoManutenzione.id.desc()).offset((page - 1) * limit).limit(limit).all()

    if not piani:
        return {"items": [], "total": 0, "page": page, "pages": 1}

    piano_ids = [p.id for p in piani]

    # Conteggio task per piano
    task_counts = dict(
        db.query(AttivitaManutenzione.piano_id, func.count(AttivitaManutenzione.id))
        .filter(AttivitaManutenzione.piano_id.in_(piano_ids))
        .group_by(AttivitaManutenzione.piano_id)
        .all()
    )

    # Conteggio ticket aperti per piano
    open_ticket_counts = dict(
        db.query(Ticket.piano_manutenzione_id, func.count(Ticket.id))
        .filter(
            Ticket.piano_manutenzione_id.in_(piano_ids),
            Ticket.stato.in_(["Aperto", "Pianificato", "In corso"]),
            Ticket.deleted_at.is_(None),
        )
        .group_by(Ticket.piano_manutenzione_id)
        .all()
    )

    # Asset IDs per piano — query esplicita sulla join table (evita lazy loading con sessione invalida)
    piano_asset_ids: dict[int, list[int]] = {pid: [] for pid in piano_ids}
    try:
        rows = (
            db.query(piano_asset_association.c.piano_id, piano_asset_association.c.asset_id)
            .filter(piano_asset_association.c.piano_id.in_(piano_ids))
            .all()
        )
        for r in rows:
            piano_asset_ids.setdefault(r[0], []).append(r[1])
    except Exception:
        db.rollback()  # reset sessione per evitare cascata di errori

    # Nome del primo asset per piano (batch — no N+1)
    all_asset_ids = {aid for ids in piano_asset_ids.values() for aid in ids}
    # Aggiungi anche gli asset_id legacy (FK diretto) per piani senza voci in join table
    for p in piani:
        if p.asset_id and not piano_asset_ids.get(p.id):
            all_asset_ids.add(p.asset_id)
    asset_name_map: dict[int, str] = {}
    if all_asset_ids:
        try:
            asset_name_map = {
                a.id: a.nome
                for a in db.query(Asset).filter(Asset.id.in_(all_asset_ids)).all()
            }
        except Exception:
            db.rollback()

    def _asset_ids_for(p: PianoManutenzione) -> list[int]:
        ids = piano_asset_ids.get(p.id, [])
        if not ids and p.asset_id:
            ids = [p.asset_id]
        return ids

    def _asset_nome_for(p: PianoManutenzione) -> str:
        ids = _asset_ids_for(p)
        return asset_name_map.get(ids[0], "") if ids else ""

    items = [
        {
            "id": p.id,
            "nome_codificato": p.nome_codificato,
            "progressivo": p.progressivo,
            "descrizione": p.descrizione,
            "stato": p.stato or "attivo",
            "asset_id": p.asset_id,
            "asset_nome": _asset_nome_for(p),
            "asset_ids": _asset_ids_for(p),
            "asset_count": len(_asset_ids_for(p)),
            "task_count": task_counts.get(p.id, 0),
            "open_ticket_count": open_ticket_counts.get(p.id, 0),
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "updated_at": p.updated_at.isoformat() if p.updated_at else None,
        }
        for p in piani
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "pages": max(1, math.ceil(total / limit)),
    }


@router.post("/piani-manutenzione", response_model=PianoManutenzioneResponse)
def create_piano(
    data: PianoManutenzioneCreate,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Crea un Piano di Manutenzione con supporto multi-asset."""
    # Gestione asset_ids (nuova) vs asset_id (legacy)
    asset_ids = data.asset_ids or []
    if not asset_ids and data.asset_id:
        asset_ids = [data.asset_id]
        
    if not asset_ids:
        raise HTTPException(status_code=400, detail="Almeno un asset_id è richiesto")

    # Verifica ownership asset
    for aid in asset_ids:
        check_tenant_ownership(db, Asset, aid, tenant_id)

    # Progressivo e Codice auto-generato (PM-2026-001)
    # Cerchiamo il max progressivo globale per questo tenant per mantenere la sequenza
    max_prog = db.query(func.max(PianoManutenzione.progressivo)).filter(
        PianoManutenzione.tenant_id == tenant_id
    ).scalar()
    prog = (max_prog or 0) + 1
    
    anno_corrente = datetime.now().year
    nome = data.nome_codificato or f"PM-{anno_corrente}-{prog:03d}"

    # Unicità forzata
    exist = db.query(PianoManutenzione).filter(
        PianoManutenzione.nome_codificato == nome,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    attempts = 0
    while exist and attempts < 100:
        prog += 1
        nome = f"PM-{anno_corrente}-{prog:03d}"
        exist = db.query(PianoManutenzione).filter(
            PianoManutenzione.nome_codificato == nome,
            PianoManutenzione.tenant_id == tenant_id,
        ).first()
        attempts += 1

    # Recupera i veri oggetti Asset
    assets_db = db.query(Asset).filter(Asset.id.in_(asset_ids)).all()

    piano = PianoManutenzione(
        tenant_id=tenant_id,
        nome_codificato=nome,
        progressivo=prog,
        descrizione=data.descrizione,
        stato=data.stato or "attivo",
        asset_id=asset_ids[0],  # legacy compat
        impianto_id=data.impianto_id,
        sito_id=data.sito_id,
        manuale_id=data.manuale_id,
    )
    piano.assets = assets_db  # relazione many-to-many
    
    db.add(piano)
    db.commit()
    db.refresh(piano)
    db_info(db, "PIANO", f"Piano {piano.nome_codificato} creato per {len(asset_ids)} asset", tenant_id=tenant_id)
    return piano


@router.get("/piani-manutenzione/{piano_id}", response_model=PianoManutenzioneResponse)
def get_piano(
    piano_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")
    
    # Arricchimento response con asset_ids
    return {
        **piano.__dict__,
        "asset_ids": [a.id for a in piano.assets]
    }


@router.put("/piani-manutenzione/{piano_id}", response_model=PianoManutenzioneResponse)
def update_piano(
    piano_id: int,
    data: PianoManutenzioneUpdate,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    # Update campi semplici
    for field in ("descrizione", "stato", "impianto_id", "sito_id"):
        val = getattr(data, field, None)
        if val is not None:
            setattr(piano, field, val)

    # Update Multi-Asset
    if data.asset_ids is not None:
        # Verifica ownership nuovi asset
        for aid in data.asset_ids:
            check_tenant_ownership(db, Asset, aid, tenant_id)
        
        assets_db = db.query(Asset).filter(Asset.id.in_(data.asset_ids)).all()
        piano.assets = assets_db
        if data.asset_ids:
            piano.asset_id = data.asset_ids[0] # legacy compat
        else:
            piano.asset_id = None

    db.commit()
    db.refresh(piano)
    return piano


@router.delete("/piani-manutenzione/{piano_id}")
def delete_piano(
    piano_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    # Scollega i ticket dal piano (non li elimina — preserva lo storico)
    db.query(Ticket).filter(
        Ticket.piano_manutenzione_id == piano_id,
        Ticket.tenant_id == tenant_id,
    ).update({"piano_manutenzione_id": None})

    db.delete(piano)
    db.commit()
    return {"deleted": piano_id}


# ── TASK sub-resource ──────────────────────────────────────────────────────────

@router.get("/piani-manutenzione/{piano_id}/tasks")
def list_piano_tasks(
    piano_id: int,
    task_stato: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Lista task appartenenti al piano."""
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    query = db.query(AttivitaManutenzione).filter(
        AttivitaManutenzione.piano_id == piano_id,
        AttivitaManutenzione.tenant_id == tenant_id,
    )
    if task_stato:
        query = query.filter(AttivitaManutenzione.task_stato == task_stato)

    total = query.count()
    tasks = query.order_by(AttivitaManutenzione.id.desc()).offset((page - 1) * limit).limit(limit).all()

    task_ids = [t.id for t in tasks]
    # Conteggio ticket aperti per task (batch — no N+1)
    open_counts = dict(
        db.query(Ticket.attivita_manutenzione_id, func.count(Ticket.id))
        .filter(
            Ticket.attivita_manutenzione_id.in_(task_ids),
            Ticket.stato.in_(["Aperto", "Pianificato", "In corso"]),
            Ticket.deleted_at.is_(None),
        )
        .group_by(Ticket.attivita_manutenzione_id)
        .all()
    )

    # Asset name (dal piano — tutti i task del piano condividono l'asset)
    asset_ids = {t.asset_id for t in tasks if t.asset_id}
    latest_readings = latest_running_hours_by_asset(db, tenant_id, asset_ids)
    asset_names = {
        a.id: a.nome
        for a in db.query(Asset).filter(Asset.id.in_(asset_ids)).all()
    } if asset_ids else {}

    return {
        "items": [
            _task_to_dict(
                t,
                open_counts.get(t.id, 0),
                asset_names.get(t.asset_id),
                latest_readings.get(t.asset_id),
            )
            for t in tasks
        ],
        "total": total,
        "page": page,
        "pages": max(1, math.ceil(total / limit)),
    }


@router.post("/piani-manutenzione/{piano_id}/tasks")
def create_piano_task(
    piano_id: int,
    data: TaskCreate,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Crea un nuovo Task all'interno del Piano."""
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    task = AttivitaManutenzione(
        piano_id=piano_id,
        asset_id=piano.asset_id,   # eredita l'asset dal piano
        nome=data.nome,
        descrizione=data.descrizione,
        frequenza_giorni=data.frequenza_giorni,
        durata_ore=data.durata_ore,
        priorita=data.priorita,
        codice=data.codice,
        origine="Manuale",
        manuale_id=None,
        tenant_id=tenant_id,
        is_repeatable=data.is_repeatable,
        generation_mode=data.generation_mode,
        generate_days_before_due=data.generate_days_before_due,
        task_stato="active",
        source_type=data.source_type,
        trigger_mode=normalize_trigger_mode(data.trigger_mode),
        condition_metric=data.condition_metric or (METRIC_RUNNING_HOURS if normalize_trigger_mode(data.trigger_mode) != "calendar" else None),
        condition_threshold_hours=data.condition_threshold_hours,
        condition_last_done_hours=data.condition_last_done_hours,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    latest = latest_running_hours_by_asset(db, tenant_id, [task.asset_id]).get(task.asset_id) if task.asset_id else None
    return _task_to_dict(task, 0, latest_reading=latest)


@router.put("/piani-manutenzione/{piano_id}/tasks/{task_id}")
def update_piano_task(
    piano_id: int,
    task_id: int,
    data: TaskUpdate,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Aggiorna un Task del Piano."""
    task = db.query(AttivitaManutenzione).filter(
        AttivitaManutenzione.id == task_id,
        AttivitaManutenzione.piano_id == piano_id,
        AttivitaManutenzione.tenant_id == tenant_id,
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task non trovato")

    for field in ("nome", "descrizione", "frequenza_giorni", "durata_ore", "priorita",
                  "codice", "is_repeatable", "generation_mode", "generate_days_before_due",
                  "task_stato", "asset_id", "trigger_mode", "condition_metric",
                  "condition_threshold_hours", "condition_last_done_hours"):
        val = getattr(data, field, None)
        if val is not None:
            setattr(task, field, val)

    task.trigger_mode = normalize_trigger_mode(task.trigger_mode)
    if task.trigger_mode == "calendar":
        task.condition_metric = None
        task.condition_threshold_hours = None
        task.condition_last_done_hours = None
    elif not task.condition_metric:
        task.condition_metric = METRIC_RUNNING_HOURS

    db.commit()
    db.refresh(task)

    open_count = db.query(func.count(Ticket.id)).filter(
        Ticket.attivita_manutenzione_id == task_id,
        Ticket.stato.in_(["Aperto", "Pianificato", "In corso"]),
        Ticket.deleted_at.is_(None),
    ).scalar() or 0
    latest = latest_running_hours_by_asset(db, tenant_id, [task.asset_id]).get(task.asset_id) if task.asset_id else None
    return _task_to_dict(task, open_count, latest_reading=latest)


@router.delete("/piani-manutenzione/{piano_id}/tasks/{task_id}")
def delete_piano_task(
    piano_id: int,
    task_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Elimina un Task (i ticket collegati vengono scollegati, non eliminati)."""
    task = db.query(AttivitaManutenzione).filter(
        AttivitaManutenzione.id == task_id,
        AttivitaManutenzione.piano_id == piano_id,
        AttivitaManutenzione.tenant_id == tenant_id,
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task non trovato")

    # Scollega i ticket dal task (preserva storico)
    db.query(Ticket).filter(
        Ticket.attivita_manutenzione_id == task_id,
        Ticket.tenant_id == tenant_id,
    ).update({"attivita_manutenzione_id": None})

    db.delete(task)
    db.commit()
    return {"deleted": task_id}


@router.post("/piani-manutenzione/{piano_id}/tasks/{task_id}/genera-ticket")
def genera_ticket_da_task(
    piano_id: int,
    task_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """
    Genera un Ticket dal Task selezionato.

    Regole:
    - Il ticket nasce sempre in stato APERTO.
    - origin_type = task_manual_generation.
    - Previene duplicati: se esiste già un ticket aperto/pianificato/in_corso per questo task
      nella stessa occorrenza, non ne crea un altro.
    - Aggiorna last_generated_at e next_due_at sul task.
    """
    task = db.query(AttivitaManutenzione).filter(
        AttivitaManutenzione.id == task_id,
        AttivitaManutenzione.piano_id == piano_id,
        AttivitaManutenzione.tenant_id == tenant_id,
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task non trovato")

    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    if getattr(task, "task_stato", "active") != "active":
        raise HTTPException(status_code=422, detail="Il task non è attivo. Attivalo prima di generare ticket.")

    if getattr(task, "generation_mode", "manual") == "disabled":
        raise HTTPException(status_code=422, detail="La generazione ticket è disabilitata per questo task.")

    # Prevenzione duplicati: esiste già un ticket aperto/pianificato/in corso per questo task?
    existing = db.query(Ticket).filter(
        Ticket.attivita_manutenzione_id == task_id,
        Ticket.stato.in_(["Aperto", "Pianificato", "In corso"]),
        Ticket.deleted_at.is_(None),
    ).first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Esiste già un ticket attivo (#{existing.id}) per questo task. Chiudi o elimina il precedente prima di generarne uno nuovo.",
        )

    durata = float(task.durata_ore) if task.durata_ore else 1.0
    titolo_base = (task.nome or task.descrizione or f"Task #{task_id}")[:150]

    # Determinazione asset target
    # Se il piano ha una lista di assets (Multi-Asset), generiamo ticket per TUTTI
    target_assets = piano.assets if piano.assets else []
    if not target_assets and task.asset_id:
        # Fallback 1: asset_id diretto sul task
        asset_legacy = db.query(Asset).filter(Asset.id == task.asset_id).first()
        if asset_legacy:
            target_assets = [asset_legacy]
    if not target_assets and piano.asset_id:
        # Fallback 2: asset_id legacy sul piano (piani creati prima della relazione M2M)
        asset_legacy = db.query(Asset).filter(Asset.id == piano.asset_id).first()
        if asset_legacy:
            target_assets = [asset_legacy]

    if not target_assets:
        raise HTTPException(status_code=422, detail="Nessun asset associato a questo piano/task.")

    created_ids = []
    for asset in target_assets:
        # Gestione ticket multi-chunk se durata > 8h
        ore_rimanenti = durata
        chunk_idx = 1
        num_chunks = int(ore_rimanenti // 8) + (1 if ore_rimanenti % 8 > 0 else 0)

        while ore_rimanenti > 0:
            durata_chunk = min(8.0, ore_rimanenti)
            titolo_chunk = f"{titolo_base} (Parte {chunk_idx}/{num_chunks})" if num_chunks > 1 else titolo_base
            if len(target_assets) > 1:
                titolo_chunk = f"{titolo_chunk} - {asset.nome}"

            ticket = Ticket(
                titolo=titolo_chunk[:150],
                descrizione=task.descrizione or "",
                asset_id=asset.id,
                tipo="PM",
                priorita=task.priorita or "Media",
                stato="Aperto",
                durata_stimata_ore=durata_chunk,
                fascia_oraria="diurna",
                attivita_manutenzione_id=task_id,
                piano_manutenzione_id=piano_id,
                origine_piano="task_manual_generation",
                origin_type="task_manual_generation",
                tenant_id=tenant_id,
            )
            db.add(ticket)
            db.flush()
            created_ids.append(ticket.id)

            ore_rimanenti -= durata_chunk
            chunk_idx += 1

    # Aggiorna metadati del task
    now = datetime.now(timezone.utc)
    task.last_generated_at = now
    if task.frequenza_giorni:
        task.next_due_at = now + timedelta(days=task.frequenza_giorni)

    db.commit()
    return {
        "created": len(created_ids),
        "ticket_ids": created_ids,
        "task_id": task_id,
        "piano_id": piano_id,
        "assets_count": len(target_assets)
    }


@router.get("/piani-manutenzione/{piano_id}/tasks/{task_id}/tickets")
def list_task_tickets(
    piano_id: int,
    task_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Storico ticket generati dal task."""
    task = db.query(AttivitaManutenzione).filter(
        AttivitaManutenzione.id == task_id,
        AttivitaManutenzione.piano_id == piano_id,
        AttivitaManutenzione.tenant_id == tenant_id,
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task non trovato")

    tickets = (
        db.query(Ticket)
        .filter(
            Ticket.attivita_manutenzione_id == task_id,
            Ticket.deleted_at.is_(None),
        )
        .order_by(Ticket.created_at.desc())
        .limit(50)
        .all()
    )
    return [
        {
            "id": t.id,
            "titolo": t.titolo,
            "stato": t.stato,
            "priorita": t.priorita,
            "durata_stimata_ore": t.durata_stimata_ore,
            "origin_type": getattr(t, "origin_type", None),
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "planned_start": t.planned_start.isoformat() if t.planned_start else None,
            "planned_finish": t.planned_finish.isoformat() if t.planned_finish else None,
            "tecnico_id": t.tecnico_id,
        }
        for t in tickets
    ]


# ── IMPORT ────────────────────────────────────────────────────────────────────

@router.post("/piani-manutenzione/{piano_id}/import-pdf")
async def import_pdf_to_piano(
    piano_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Importa attività da PDF (AI/OCR) come TASK del piano."""
    from backend.services.pdf_service import smart_read_pdf
    from backend.services.ai.manuals_ai_service import parse_manual_with_ai
    import json

    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File vuoto.")
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File troppo grande: massimo {MAX_IMPORT_BYTES // (1024 * 1024)} MB consentiti.",
        )

    result = smart_read_pdf(content)
    text = result.get("text", "")

    if not text.strip():
        raise HTTPException(status_code=422, detail="Nessun testo estratto dal PDF.")

    parsed_json_str = parse_manual_with_ai(text, file.filename)
    try:
        parsed = json.loads(parsed_json_str)
    except Exception:
        raise HTTPException(status_code=500, detail="L'AI non ha restituito un formato valido.")

    plans = parsed.get("plans", [])
    created_count = 0

    for plan in plans:
        freq_days = _freq_days(plan.get("frequency"))
        durata = plan.get("estimated_duration_hours") or 1.0
        priorita = _PRIO_MAP.get((plan.get("priority") or "medium").lower(), "Media")

        for task_str in plan.get("tasks", []):
            if not task_str:
                continue
            task = AttivitaManutenzione(
                piano_id=piano_id,
                asset_id=piano.asset_id,
                nome=task_str[:150],
                descrizione=task_str,
                priorita=priorita,
                frequenza_giorni=freq_days,
                durata_ore=durata,
                origine="PDF",
                tenant_id=tenant_id,
                task_stato="active",
                generation_mode="manual",
                is_repeatable=True,
                source_type="imported_from_manual",
            )
            db.add(task)
            created_count += 1

    db.commit()
    return {"success": True, "created": created_count, "method": result.get("method")}



@router.post("/piani-manutenzione/{piano_id}/import-excel")
async def import_excel_to_piano(
    piano_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Importa attività da Excel/CSV come TASK del piano (non ticket diretti)."""
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File vuoto.")
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File troppo grande: massimo {MAX_IMPORT_BYTES // (1024 * 1024)} MB consentiti.",
        )

    filename = (file.filename or "").lower()

    rows = []
    if filename.endswith(".csv"):
        stream = io.StringIO(content.decode("utf-8", errors="ignore"))
        reader = csv.DictReader(stream, delimiter=";")
        rows = list(reader)
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            header = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
            for i in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, col_name in enumerate(header):
                    if col_name:
                        row_data[col_name] = ws.cell(row=i, column=col_idx + 1).value
                if any(row_data.values()):
                    rows.append(row_data)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {e}")

    created_count = 0
    for r in rows:
        nome = r.get("nome") or r.get("titolo") or r.get("attività") or r.get("descrizione")
        if not nome:
            continue
        durata_raw = r.get("durata") or r.get("ore") or 1.0
        priorita_raw = r.get("priorità") or r.get("priorita") or "Media"
        freq_raw = r.get("frequenza") or r.get("frequenza_giorni")

        task = AttivitaManutenzione(
            piano_id=piano_id,
            asset_id=piano.asset_id,
            nome=str(nome)[:150],
            descrizione=str(nome),
            priorita=str(priorita_raw).capitalize() if priorita_raw else "Media",
            durata_ore=float(durata_raw) if durata_raw else 1.0,
            frequenza_giorni=int(freq_raw) if freq_raw else None,
            origine="Excel",
            tenant_id=tenant_id,
            task_stato="active",
            generation_mode="manual",
            is_repeatable=True,
            source_type="imported_from_manual",
        )
        db.add(task)
        created_count += 1

    db.commit()
    return {"success": True, "created": created_count}


# ── BACKWARD COMPAT: ticket assignment (legacy) ───────────────────────────────

@router.post("/piani-manutenzione/{piano_id}/ticket")
def assign_ticket_to_piano(
    piano_id: int,
    ticket_id: int,
    origine: str = Query("manuale_interno_piano"),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Collega un ticket esistente al piano (backward compat)."""
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    ticket = db.query(Ticket).filter(
        Ticket.id == ticket_id,
        Ticket.tenant_id == tenant_id,
    ).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket non trovato")

    ticket.piano_manutenzione_id = piano_id
    ticket.origine_piano = origine
    db.commit()
    return {"success": True, "ticket_id": ticket_id, "piano_id": piano_id}


# ── MANUALI (Integrated) ───────────────────────────────────────────────────

@router.get("/piani-manutenzione/{piano_id}/manuali")
def list_piano_manuali(
    piano_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Lista manuali PDF associati al piano."""
    piano = db.query(PianoManutenzione).filter(
        PianoManutenzione.id == piano_id,
        PianoManutenzione.tenant_id == tenant_id,
    ).first()
    if not piano:
        raise HTTPException(status_code=404, detail="Piano non trovato")

    manuali = db.query(Manuale).filter(
        Manuale.piano_id == piano_id,
        Manuale.tenant_id == tenant_id,
    ).all()
    
    return [
        {
            "id": m.id,
            "nome": m.nome_file,
            "pagine": m.pagine,
            "metodo": m.metodo_lettura,
            "created_at": m.created_at.isoformat() if hasattr(m, "created_at") and m.created_at else None,
            "stato": m.stato or "attivo"
        }
        for m in manuali
    ]

@router.delete("/piani-manutenzione/{piano_id}/manuali/{manuale_id}")
def delete_piano_manuale(
    piano_id: int,
    manuale_id: int,
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Scollega o elimina un manuale dal piano."""
    manuale = db.query(Manuale).filter(
        Manuale.id == manuale_id,
        Manuale.piano_id == piano_id,
        Manuale.tenant_id == tenant_id,
    ).first()
    if not manuale:
        raise HTTPException(status_code=404, detail="Manuale non trovato nel contesto di questo piano")
    
    db.delete(manuale)
    db.commit()
    return {"deleted": manuale_id}
