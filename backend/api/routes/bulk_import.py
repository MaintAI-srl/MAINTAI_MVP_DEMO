"""
Bulk Import — Caricamento massivo Siti / Impianti / Asset da Excel.
Riservato esclusivamente ai superadmin.

Endpoint:
  GET  /admin/bulk-import/template         → scarica template Excel con esempio + colonne gialle
  POST /admin/bulk-import/preview          → carica file Excel, valida e restituisce anteprima
  POST /admin/bulk-import/execute/{token}  → esegue l'import per un tenant specifico

Struttura Excel (3 fogli):
  1. SITI     — Colonne obbligatorie gialle: nome
  2. IMPIANTI — Colonne obbligatorie gialle: nome, sito_nome
  3. ASSET    — Colonne obbligatorie gialle: nome, area, impianto_nome
"""
from __future__ import annotations

import io
import json
import logging
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.core.dependencies import get_db
from backend.core.security import require_superadmin
from backend.core.file_validation import validate_upload
from backend.db.modelli import Asset, Attestato, AttivitaManutenzione, Impianto, PianoManutenzione, Sito, Tecnico

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin/bulk-import", tags=["bulk-import"])

MAX_IMPORT_BYTES = 10 * 1024 * 1024  # 10 MB — limite upload Excel (anti memory-exhaustion)

# ── Definizione colonne per foglio ────────────────────────────────────────────

# (nome_colonna, obbligatoria, descrizione, esempio)
SITI_COLS: List[tuple] = [
    ("nome",                    True,  "Nome del sito (obbligatorio)",           "Porto Industriale Nord"),
    ("descrizione",             False, "Descrizione estesa del sito",             "Sito produttivo principale"),
    ("ubicazione",              False, "Indirizzo / ubicazione",                  "Via Industriale 42"),
    ("citta",                   False, "Città",                                   "Genova"),
    ("paese",                   False, "Paese (default: Italia)",                 "Italia"),
    ("responsabile",            False, "Nome responsabile di sito",               "Mario Rossi"),
    ("telefono_responsabile",   False, "Telefono del responsabile",               "+39 010 1234567"),
    ("email_responsabile",      False, "Email del responsabile",                  "m.rossi@azienda.it"),
    ("note",                    False, "Note libere",                             "Apertura 6:00-22:00"),
]

IMPIANTI_COLS: List[tuple] = [
    ("nome",        True,  "Nome impianto (obbligatorio)",                        "Linea Produzione A"),
    ("sito_nome",   True,  "Nome del sito padre (deve esistere nel foglio SITI)", "Porto Industriale Nord"),
    ("tipologia",   False, "Tipologia impianto (es. Produzione, Energia…)",       "Produzione"),
    ("descrizione", False, "Descrizione estesa",                                  "Linea confezionamento"),
    ("latitude",    False, "Latitudine GPS (per meteo)",                          "44.4056"),
    ("longitude",   False, "Longitudine GPS (per meteo)",                         "8.9463"),
    ("note",        False, "Note libere",                                          "Manutenzione ogni 6 mesi"),
]

ASSET_COLS: List[tuple] = [
    ("nome",                   True,  "Nome asset (obbligatorio)",                          "Compressore C-01"),
    ("area",                   True,  "Area / zona di appartenenza (obbligatorio)",         "Area Compressori"),
    ("impianto_nome",          True,  "Nome dell'impianto padre (deve esistere in IMPIANTI)", "Linea Produzione A"),
    ("codice",                 False, "Codice identificativo asset",                        "COMP-001"),
    ("descrizione",            False, "Descrizione tecnica",                                "Compressore a vite 37kW"),
    ("criticita",              False, "Criticità: bassa / media / alta (default: media)",   "alta"),
    ("stato",                  False, "Stato asset: OPERATIVO / FERMO PROG. / GUASTO",      "OPERATIVO"),
    ("marca",                  False, "Marca / costruttore",                                "Atlas Copco"),
    ("modello",                False, "Modello",                                            "GA37+"),
    ("matricola",              False, "Matricola",                                          "AC-2024-001"),
    ("numero_serie",           False, "Numero di serie",                                   "SN123456789"),
    ("anno_installazione",     False, "Anno di installazione (numerico)",                  "2020"),
    ("anno_produzione",        False, "Anno di produzione (numerico)",                     "2019"),
    ("fornitore",              False, "Nome fornitore/distributore",                        "Atlas Copco Italia"),
    ("data_acquisto",          False, "Data acquisto (YYYY-MM-DD)",                        "2020-03-15"),
    ("data_scadenza_garanzia", False, "Data scadenza garanzia (YYYY-MM-DD)",               "2025-03-15"),
    ("posizione_fisica",       False, "Posizione fisica nell'impianto",                    "Sala macchine piano -1"),
    ("limitazioni",            False, "Limitazioni operative",                             "no_notturno"),
    ("weather_constraint",     False, "Vincolo meteo: NONE/NO_RAIN/NO_WIND/NO_FROST",      "NONE"),
    ("note",                   False, "Note operative",                                    "Revisione ogni 2000h"),
    ("note_tecniche",          False, "Note tecniche dettagliate",                         "Olio Shell Corena"),
    ("vincoli_operativi",      False, "Vincoli operativi",                                 "Fermare prima della manutenzione"),
    ("vincoli_manutenzione",   False, "Vincoli specifici manutenzione",                    "Spegnere 2h prima"),
    ("fermo_on_schedule",      False, "Metti in FERMO PROG. al piano: SI/NO (default: NO)", "NO"),
]

TECNICI_COLS: List[tuple] = [
    ("nome", True, "Nome tecnico", "Luca"),
    ("cognome", False, "Cognome tecnico", "Bianchi"),
    ("skill", False, "Competenze separate da virgola", "Meccanico, Elettricista"),
    ("ore_giornaliere", False, "Ore giornaliere disponibili", "8"),
    ("stato", False, "in servizio / ferie / malattia / corso", "in servizio"),
    ("orario_inizio", False, "Ora inizio turno HH:MM", "08:00"),
    ("orario_fine", False, "Ora fine turno HH:MM", "17:00"),
    ("telefono", False, "Telefono", "+39 333 1234567"),
    ("sede_indirizzo", False, "Sede o indirizzo di partenza", "Via Officina 10, Genova"),
    ("limitazioni_orarie", False, "Eventuali limitazioni", "no notturno"),
]

PIANI_COLS: List[tuple] = [
    ("piano_nome", True, "Codice/nome piano da creare o aggiornare", "PM-COMP-001"),
    ("piano_descrizione", False, "Descrizione del piano", "Piano compressore linea A"),
    ("asset_id", False, "ID asset esistente, se noto", "12"),
    ("asset_codice", False, "Codice asset esistente", "COMP-001"),
    ("asset_nome", False, "Nome asset esistente se non usi asset_id", "Compressore C-01"),
    ("task_nome", True, "Nome sintetico attivita", "Cambio filtri aria"),
    ("task_descrizione", True, "Descrizione operativa", "Sostituire filtri e verificare pressione"),
    ("frequenza_giorni", False, "Frequenza in giorni", "90"),
    ("durata_ore", False, "Durata stimata in ore", "2"),
    ("priorita", False, "Alta / Media / Bassa", "Media"),
    ("generation_mode", False, "manual / auto / disabled", "manual"),
    ("generate_days_before_due", False, "Giorni anticipo generazione ticket", "7"),
]

CORSI_COLS: List[tuple] = [
    ("tecnico_id", False, "ID tecnico esistente, se noto", "3"),
    ("tecnico_nome", False, "Nome tecnico se non usi tecnico_id", "Luca"),
    ("tecnico_cognome", False, "Cognome tecnico", "Bianchi"),
    ("tipo_corso", True, "Tipo corso o attestato", "PES/PAV"),
    ("ente_certificatore", False, "Ente certificatore", "CEI"),
    ("data_conseguimento", False, "Data conseguimento YYYY-MM-DD", "2026-01-15"),
    ("data_scadenza", False, "Data scadenza YYYY-MM-DD", "2028-01-15"),
    ("note", False, "Note", "Aggiornamento biennale"),
]

SINGLE_IMPORTS: dict[str, dict[str, Any]] = {
    "siti": {"label": "Siti", "sheet": "SITI", "cols": SITI_COLS, "filename": "maintai_template_siti.xlsx"},
    "impianti": {"label": "Impianti", "sheet": "IMPIANTI", "cols": IMPIANTI_COLS, "filename": "maintai_template_impianti.xlsx"},
    "asset": {"label": "Asset", "sheet": "ASSET", "cols": ASSET_COLS, "filename": "maintai_template_asset.xlsx"},
    "tecnici": {"label": "Tecnici", "sheet": "TECNICI", "cols": TECNICI_COLS, "filename": "maintai_template_tecnici.xlsx"},
    "piani_manutenzione": {"label": "Piani di manutenzione", "sheet": "PIANI", "cols": PIANI_COLS, "filename": "maintai_template_piani_manutenzione.xlsx"},
    "corsi_attestati": {"label": "Corsi e attestati", "sheet": "CORSI", "cols": CORSI_COLS, "filename": "maintai_template_corsi_attestati.xlsx"},
}

HIERARCHY_TYPES = {"gerarchia", "siti_impianti_asset", "siti-impianti-asset", "hierarchy"}


# ── Stili Excel ───────────────────────────────────────────────────────────────

def _make_workbook() -> Any:
    """Crea un workbook openpyxl con il template compilato."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Colori
    YELLOW_FILL    = PatternFill("solid", fgColor="FFD700")   # oro/giallo — obbligatorio
    GREY_FILL      = PatternFill("solid", fgColor="D9D9D9")   # grigio chiaro — opzionale
    HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")   # blu scuro — intestazione
    EXAMPLE_FILL   = PatternFill("solid", fgColor="EBF3FB")   # azzurro chiaro — riga esempio
    LEGEND_REQ     = PatternFill("solid", fgColor="FFD700")
    LEGEND_OPT     = PatternFill("solid", fgColor="D9D9D9")

    WHITE_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BOLD_FONT      = Font(bold=True, name="Calibri", size=10)
    NORMAL_FONT    = Font(name="Calibri", size=10)
    SMALL_FONT     = Font(name="Calibri", size=9, italic=True, color="595959")

    thin_border    = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def _setup_sheet(ws, title: str, cols: List[tuple], sheet_color: str):
        """Configura un foglio con header, legenda e riga esempio."""
        ws.title = title
        ws.sheet_properties.tabColor = sheet_color

        # ── Riga 1: Titolo foglio ────────────────────────────────────────────
        ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        title_cell = ws["A1"]
        title_cell.value = f"  {title} — Template importazione MaintAI"
        title_cell.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        title_cell.fill = PatternFill("solid", fgColor="0D2137")
        title_cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[1].height = 28

        # ── Riga 2: Legenda ──────────────────────────────────────────────────
        ws.merge_cells(f"A2:{get_column_letter(len(cols))}2")
        leg_cell = ws["A2"]
        leg_cell.value = "  GIALLO = campo obbligatorio    GRIGIO = campo opzionale    La riga 4 è un esempio: modificarla o eliminarla prima dell'import"
        leg_cell.font = Font(size=9, italic=True, color="404040", name="Calibri")
        leg_cell.fill = PatternFill("solid", fgColor="F2F2F2")
        leg_cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[2].height = 18

        # ── Riga 3: Header colonne ───────────────────────────────────────────
        ws.row_dimensions[3].height = 22
        for col_idx, (col_name, required, desc, example) in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name.upper())
            cell.font = WHITE_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # Colore colonna nel body: giallo o grigio
            col_fill = YELLOW_FILL if required else GREY_FILL
            # Header indicatore obbligatorietà — asterisco per required
            if required:
                cell.value = f"* {col_name.upper()}"

            # Larghezza colonna
            ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(col_name) + 4)

            # Commento colonna (descrizione) tramite Note
            # openpyxl non supporta commenti nativi senza openpyxl-image, usiamo riga desc
            # Riga 4: riga descrizione colonna (legenda interna)
            desc_cell = ws.cell(row=4, column=col_idx, value=desc)
            desc_cell.font = SMALL_FONT
            desc_cell.fill = PatternFill("solid", fgColor="FAFAFA")
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            desc_cell.border = thin_border

        ws.row_dimensions[4].height = 30

        # ── Riga 5: Header "colore" (obbligatorio/opzionale visivo) ──────────
        ws.row_dimensions[5].height = 18
        for col_idx, (col_name, required, desc, example) in enumerate(cols, start=1):
            marker_cell = ws.cell(row=5, column=col_idx, value="▼ obbligatorio" if required else "opzionale")
            marker_cell.font = Font(size=8, bold=required, name="Calibri",
                                    color="5C3317" if required else "595959")
            marker_cell.fill = YELLOW_FILL if required else GREY_FILL
            marker_cell.alignment = Alignment(horizontal="center")
            marker_cell.border = thin_border

        # ── Riga 6: Riga esempio ─────────────────────────────────────────────
        ws.row_dimensions[6].height = 20
        for col_idx, (col_name, required, desc, example) in enumerate(cols, start=1):
            ex_cell = ws.cell(row=6, column=col_idx, value=example)
            ex_cell.font = NORMAL_FONT
            ex_cell.fill = EXAMPLE_FILL
            ex_cell.alignment = Alignment(vertical="center")
            ex_cell.border = thin_border

        # ── Riga 7+: righe vuote per inserimento dati ────────────────────────
        for row in range(7, 207):  # 200 righe vuote
            for col_idx, (col_name, required, _, __) in enumerate(cols, start=1):
                cell = ws.cell(row=row, column=col_idx, value=None)
                cell.fill = YELLOW_FILL if required else PatternFill("solid", fgColor="FFFFFF")
                cell.font = NORMAL_FONT
                cell.border = Border(
                    left=Side(style="thin", color="E0E0E0"),
                    right=Side(style="thin", color="E0E0E0"),
                    top=Side(style="dotted", color="E0E0E0"),
                    bottom=Side(style="dotted", color="E0E0E0"),
                )
                cell.alignment = Alignment(vertical="center")

        # Freeze panes sotto l'header
        ws.freeze_panes = "A7"

        # Filtro automatico sugli header
        ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{206}"

    # Configura i tre fogli
    ws_siti = wb.active
    _setup_sheet(ws_siti, "SITI", SITI_COLS, "1F78D1")

    ws_impianti = wb.create_sheet()
    _setup_sheet(ws_impianti, "IMPIANTI", IMPIANTI_COLS, "28A745")

    ws_asset = wb.create_sheet()
    _setup_sheet(ws_asset, "ASSET", ASSET_COLS, "E67E22")

    # ── Foglio ISTRUZIONI ────────────────────────────────────────────────────
    ws_help = wb.create_sheet("ISTRUZIONI")
    ws_help.sheet_properties.tabColor = "7F7F7F"
    instructions = [
        ("", ""),
        ("  GUIDA IMPORTAZIONE MASSIVA — MaintAI", ""),
        ("", ""),
        ("  COME USARE QUESTO FILE:", ""),
        ("  1.", "Compila il foglio SITI con i siti da importare"),
        ("  2.", "Compila il foglio IMPIANTI — la colonna 'sito_nome' DEVE corrispondere esattamente al campo 'nome' nel foglio SITI"),
        ("  3.", "Compila il foglio ASSET — la colonna 'impianto_nome' DEVE corrispondere esattamente al campo 'nome' nel foglio IMPIANTI"),
        ("  4.", "I campi con asterisco (*) sono OBBLIGATORI — le righe con valori mancanti verranno rifiutate"),
        ("  5.", "Elimina la riga di esempio (riga 6) prima di caricare, oppure lasciala — viene ignorata se è uguale all'esempio"),
        ("  6.", "Carica il file nella pagina Admin → Import Massivo"),
        ("", ""),
        ("  VALORI AMMESSI:", ""),
        ("  criticita",  "bassa / media / alta  (default: media se vuoto)"),
        ("  stato",      "OPERATIVO / FERMO PROG. / GUASTO  (default: OPERATIVO se vuoto)"),
        ("  weather_constraint", "NONE / NO_RAIN / NO_WIND / NO_FROST / OUTDOOR_ONLY / INDOOR_ONLY  (default: NONE)"),
        ("  fermo_on_schedule", "SI / NO  (default: NO se vuoto)"),
        ("  paese",      "qualsiasi stringa  (default: Italia se vuoto)"),
        ("  date",       "formato YYYY-MM-DD, es. 2024-03-15"),
        ("  latitude / longitude", "decimali con punto, es. 44.4056 / 8.9463"),
        ("", ""),
        ("  ORDINE DI IMPORT:", ""),
        ("  →", "Prima vengono creati i SITI"),
        ("  →", "Poi gli IMPIANTI (collegati ai siti per nome)"),
        ("  →", "Infine gli ASSET (collegati agli impianti per nome)"),
        ("", ""),
        ("  ATTENZIONE:", ""),
        ("  !", "Il nome del sito e dell'impianto deve essere ESATTO (case-insensitive)"),
        ("  !", "Se un sito o impianto con lo stesso nome esiste già, NON viene duplicato — viene usato quello esistente"),
        ("  !", "L'import è reversibile: contattare il superadmin per annullare"),
    ]
    for r, (col_a, col_b) in enumerate(instructions, start=1):
        ca = ws_help.cell(row=r, column=1, value=col_a)
        cb = ws_help.cell(row=r, column=2, value=col_b)
        if r == 2:
            ca.font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
        elif col_a.strip().startswith(("COME", "VALORI", "ORDINE", "ATTENZIONE")):
            ca.font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
        else:
            ca.font = Font(size=10, name="Calibri")
            cb.font = Font(size=10, name="Calibri")
    ws_help.column_dimensions["A"].width = 28
    ws_help.column_dimensions["B"].width = 80

    return wb


def _make_single_workbook(sheet_name: str, cols: List[tuple], title: str) -> Any:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    yellow_fill = PatternFill("solid", fgColor="FFD700")
    grey_fill = PatternFill("solid", fgColor="D9D9D9")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    example_fill = PatternFill("solid", fgColor="EBF3FB")
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws["A1"] = f"  {title} - Template importazione MaintAI"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="0D2137")
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(len(cols))}2")
    ws["A2"] = "  GIALLO = obbligatorio    GRIGIO = opzionale    La riga 6 e un esempio: modificarla o eliminarla"
    ws["A2"].font = Font(size=9, italic=True, color="404040", name="Calibri")
    ws["A2"].fill = PatternFill("solid", fgColor="F2F2F2")

    for col_idx, (col_name, required, desc, example) in enumerate(cols, start=1):
        header = ws.cell(row=3, column=col_idx, value=f"* {col_name.upper()}" if required else col_name.upper())
        header.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header.fill = header_fill
        header.alignment = Alignment(horizontal="center", vertical="center")
        header.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(col_name) + 5)

        desc_cell = ws.cell(row=4, column=col_idx, value=desc)
        desc_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
        desc_cell.fill = PatternFill("solid", fgColor="FAFAFA")
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.border = thin_border

        marker = ws.cell(row=5, column=col_idx, value="obbligatorio" if required else "opzionale")
        marker.font = Font(size=8, bold=required, name="Calibri")
        marker.fill = yellow_fill if required else grey_fill
        marker.alignment = Alignment(horizontal="center")
        marker.border = thin_border

        sample = ws.cell(row=6, column=col_idx, value=example)
        sample.font = Font(name="Calibri", size=10)
        sample.fill = example_fill
        sample.border = thin_border

    for row in range(7, 207):
        for col_idx, (_, required, _, __) in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=col_idx, value=None)
            cell.fill = yellow_fill if required else PatternFill("solid", fgColor="FFFFFF")
            cell.border = Border(
                left=Side(style="thin", color="E0E0E0"),
                right=Side(style="thin", color="E0E0E0"),
                top=Side(style="dotted", color="E0E0E0"),
                bottom=Side(style="dotted", color="E0E0E0"),
            )

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}206"

    help_ws = wb.create_sheet("ISTRUZIONI")
    help_ws["A1"] = "Compila il foglio dati e carica il file dalla pagina Admin > Import Massivo."
    help_ws["A2"] = "Le colonne con asterisco sono obbligatorie. Mantieni invariati i nomi delle colonne."
    help_ws.column_dimensions["A"].width = 100
    return wb


# ── Endpoint: download template ───────────────────────────────────────────────

@router.get("/template")
def download_template(
    _sa: dict = Depends(require_superadmin),
):
    """
    Scarica il template Excel per l'importazione massiva di Siti, Impianti e Asset.
    Colonne gialle = obbligatorie. Colonne grigie = opzionali. Riga 6 = esempio.
    """
    wb = _make_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "maintai_bulk_import_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/templates")
def list_templates(_sa: dict = Depends(require_superadmin)):
    templates = [
        {
            "id": "gerarchia",
            "label": "Siti + Impianti + Asset",
            "description": "Template completo con tre fogli collegati gerarchicamente.",
            "filename": "maintai_bulk_import_template.xlsx",
        }
    ]
    templates.extend(
        {
            "id": import_type,
            "label": meta["label"],
            "description": f"Template Excel dedicato a {str(meta['label']).lower()}.",
            "filename": meta["filename"],
        }
        for import_type, meta in SINGLE_IMPORTS.items()
    )
    return {"templates": templates}


@router.get("/template/{import_type}")
def download_template_by_type(
    import_type: str,
    _sa: dict = Depends(require_superadmin),
):
    normalized = import_type.strip().lower()
    if normalized in HIERARCHY_TYPES:
        return download_template(_sa)
    meta = SINGLE_IMPORTS.get(normalized)
    if not meta:
        raise HTTPException(404, "Tipologia template non supportata")

    wb = _make_single_workbook(meta["sheet"], meta["cols"], meta["label"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{meta["filename"]}"'},
    )


# ── Parsing Excel ─────────────────────────────────────────────────────────────

_EXAMPLE_NAMES = {
    "Porto Industriale Nord", "Linea Produzione A", "Compressore C-01",
}

def _cell_val(row, col_name: str, col_map: Dict[str, int]) -> Optional[str]:
    """Estrae il valore di una cella per nome colonna (strip + None se vuoto)."""
    idx = col_map.get(col_name)
    if idx is None:
        return None
    val = row[idx].value
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_sheet(ws, expected_cols: List[tuple]) -> tuple[List[Dict], List[str]]:
    """
    Legge un foglio Excel e restituisce (righe_valide, errori).
    Ignora le prime 5 righe (titolo, legenda, header, descrizione, marker).
    Riga 6 = esempio — viene ignorata se il nome corrisponde ai nomi di esempio.
    """
    required = {c[0] for c in expected_cols if c[1]}
    col_names = [c[0] for c in expected_cols]

    # Leggi header da riga 3 (row index 3, 1-based)
    header_row = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    col_map: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h:
            # Normalizza: rimuovi asterisco e spazi, lowercase
            key = str(h).replace("*", "").strip().lower()
            col_map[key] = i

    rows: List[Dict] = []
    errors: List[str] = []

    for row_num in range(6, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c) for c in range(1, ws.max_column + 1)]

        # Skip righe completamente vuote
        if all(c.value is None or str(c.value).strip() == "" for c in row):
            continue

        record: Dict[str, Any] = {}
        row_errors: List[str] = []

        for col_name in col_names:
            val = _cell_val(row, col_name, col_map)
            record[col_name] = val

        # Skip riga esempio
        nome = record.get("nome", "") or ""
        if nome in _EXAMPLE_NAMES:
            continue
        if row_num == 6:
            example_values = {col_name: str(example).strip() for col_name, _, _, example in expected_cols}
            non_empty = {key: str(value).strip() for key, value in record.items() if value is not None}
            if non_empty and all(non_empty.get(key) == example_values.get(key) for key in non_empty):
                continue

        # Verifica obbligatori
        for req in required:
            if not record.get(req):
                row_errors.append(f"Riga {row_num}: campo obbligatorio '{req}' mancante")

        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append({"_row": row_num, **record})

    return rows, errors


# ── Endpoint: preview ─────────────────────────────────────────────────────────

@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    tenant_id: int = Form(...),
    import_type: str = Form("gerarchia"),
    _sa: dict = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    """
    Carica un file Excel e restituisce un'anteprima dell'import senza scrivere nel DB.
    Valida i riferimenti tra SITI → IMPIANTI → ASSET.
    Parametri:
      - file: file .xlsx
      - tenant_id: ID del tenant destinatario dell'import
    """
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Il file deve essere in formato .xlsx")

    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(413, f"File troppo grande: massimo {MAX_IMPORT_BYTES // (1024 * 1024)} MB.")

    # Validazione magic bytes: xlsx è famiglia ZIP (PK\x03\x04)
    try:
        validate_upload(content, file.filename, {"xlsx"})
    except ValueError as exc:
        raise HTTPException(400, f"File Excel non valido: {exc}")

    normalized_type = import_type.strip().lower()
    if normalized_type not in HIERARCHY_TYPES:
        return _preview_single_import(normalized_type, content, tenant_id, db)

    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Impossibile leggere il file Excel: {e}")

    # Leggi i tre fogli
    ws_siti     = wb["SITI"]     if "SITI"     in wb.sheetnames else None
    ws_impianti = wb["IMPIANTI"] if "IMPIANTI" in wb.sheetnames else None
    ws_asset    = wb["ASSET"]    if "ASSET"    in wb.sheetnames else None

    if not ws_siti or not ws_impianti or not ws_asset:
        raise HTTPException(400, "Il file deve contenere i fogli: SITI, IMPIANTI, ASSET")

    all_errors: List[str] = []

    siti_rows, siti_errs = _parse_sheet(ws_siti, SITI_COLS)
    impianti_rows, imp_errs = _parse_sheet(ws_impianti, IMPIANTI_COLS)
    asset_rows, asset_errs = _parse_sheet(ws_asset, ASSET_COLS)

    all_errors.extend(siti_errs)
    all_errors.extend(imp_errs)
    all_errors.extend(asset_errs)

    # Verifica cross-references
    siti_nomi = {r["nome"].strip().lower() for r in siti_rows if r.get("nome")}
    # Aggiungi anche i siti già esistenti nel DB per il tenant
    siti_db = {
        s.nome.strip().lower()
        for s in db.query(Sito).filter(Sito.tenant_id == tenant_id).all()
    }
    siti_nomi_totali = siti_nomi | siti_db

    impianti_nomi = {r["nome"].strip().lower() for r in impianti_rows if r.get("nome")}
    impianti_db = {
        i.nome.strip().lower()
        for i in db.query(Impianto).filter(Impianto.tenant_id == tenant_id).all()
    }
    impianti_nomi_totali = impianti_nomi | impianti_db

    for r in impianti_rows:
        sito_ref = (r.get("sito_nome") or "").strip().lower()
        if sito_ref and sito_ref not in siti_nomi_totali:
            all_errors.append(
                f"Foglio IMPIANTI, riga {r['_row']}: sito_nome '{r['sito_nome']}' non trovato in SITI né nel DB"
            )

    for r in asset_rows:
        imp_ref = (r.get("impianto_nome") or "").strip().lower()
        if imp_ref and imp_ref not in impianti_nomi_totali:
            all_errors.append(
                f"Foglio ASSET, riga {r['_row']}: impianto_nome '{r['impianto_nome']}' non trovato in IMPIANTI né nel DB"
            )

    return {
        "valid": len(all_errors) == 0,
        "errors": all_errors,
        "counts": {
            "siti":     len(siti_rows),
            "impianti": len(impianti_rows),
            "asset":    len(asset_rows),
        },
        "preview": {
            "siti":     [_clean_row(r) for r in siti_rows[:5]],
            "impianti": [_clean_row(r) for r in impianti_rows[:5]],
            "asset":    [_clean_row(r) for r in asset_rows[:5]],
        },
    }


def _clean_row(r: Dict) -> Dict:
    """Rimuove la chiave interna _row per la serializzazione."""
    return {k: v for k, v in r.items() if k != "_row"}


def _find_sito(db: Session, tenant_id: int, nome: str | None) -> Sito | None:
    if not nome:
        return None
    return db.query(Sito).filter(Sito.tenant_id == tenant_id, Sito.nome.ilike(nome.strip())).first()


def _find_impianto(db: Session, tenant_id: int, nome: str | None) -> Impianto | None:
    if not nome:
        return None
    return db.query(Impianto).filter(Impianto.tenant_id == tenant_id, Impianto.nome.ilike(nome.strip())).first()


def _find_asset(db: Session, tenant_id: int, row: Dict[str, Any]) -> Asset | None:
    asset_id = _to_int(row.get("asset_id"))
    if asset_id:
        asset = db.query(Asset).filter(Asset.id == asset_id, Asset.tenant_id == tenant_id).first()
        if asset:
            return asset
    codice = row.get("asset_codice")
    if codice:
        asset = db.query(Asset).filter(Asset.tenant_id == tenant_id, Asset.codice.ilike(codice.strip())).first()
        if asset:
            return asset
    nome = row.get("asset_nome") or row.get("nome")
    if nome:
        return db.query(Asset).filter(Asset.tenant_id == tenant_id, Asset.nome.ilike(nome.strip())).first()
    return None


def _find_tecnico(db: Session, tenant_id: int, row: Dict[str, Any]) -> Tecnico | None:
    tecnico_id = _to_int(row.get("tecnico_id"))
    if tecnico_id:
        tecnico = db.query(Tecnico).filter(Tecnico.id == tecnico_id, Tecnico.tenant_id == tenant_id).first()
        if tecnico:
            return tecnico
    nome = (row.get("tecnico_nome") or row.get("nome") or "").strip()
    cognome = (row.get("tecnico_cognome") or row.get("cognome") or "").strip()
    if not nome:
        return None
    query = db.query(Tecnico).filter(Tecnico.tenant_id == tenant_id, Tecnico.nome.ilike(nome))
    if cognome:
        query = query.filter(Tecnico.cognome.ilike(cognome))
    return query.first()


def _read_single_rows(content: bytes, import_type: str) -> tuple[List[Dict], List[str]]:
    from openpyxl import load_workbook

    meta = SINGLE_IMPORTS.get(import_type)
    if not meta:
        raise HTTPException(400, "Tipologia import non supportata")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Impossibile leggere il file Excel: {exc}")

    ws = wb[meta["sheet"]] if meta["sheet"] in wb.sheetnames else wb.active
    return _parse_sheet(ws, meta["cols"])


def _validate_single_rows(import_type: str, rows: List[Dict], db: Session, tenant_id: int) -> List[str]:
    errors: List[str] = []
    for row in rows:
        row_num = row.get("_row", "?")
        if import_type == "impianti" and not _find_sito(db, tenant_id, row.get("sito_nome")):
            errors.append(f"Riga {row_num}: sito_nome '{row.get('sito_nome')}' non trovato nel tenant")
        elif import_type == "asset" and not _find_impianto(db, tenant_id, row.get("impianto_nome")):
            errors.append(f"Riga {row_num}: impianto_nome '{row.get('impianto_nome')}' non trovato nel tenant")
        elif import_type == "tecnici":
            stato = (row.get("stato") or "in servizio").strip().lower()
            if stato not in {"in servizio", "ferie", "malattia", "corso"}:
                errors.append(f"Riga {row_num}: stato tecnico non valido")
        elif import_type == "piani_manutenzione" and not _find_asset(db, tenant_id, row):
            errors.append(f"Riga {row_num}: asset non trovato (usa asset_id, asset_codice o asset_nome esistente)")
        elif import_type == "corsi_attestati" and not _find_tecnico(db, tenant_id, row):
            errors.append(f"Riga {row_num}: tecnico non trovato (usa tecnico_id o nome/cognome esistente)")
    return errors


def _preview_single_import(import_type: str, content: bytes, tenant_id: int, db: Session) -> dict:
    rows, errors = _read_single_rows(content, import_type)
    errors.extend(_validate_single_rows(import_type, rows, db, tenant_id))
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "counts": {import_type: len(rows)},
        "preview": {import_type: [_clean_row(r) for r in rows[:5]]},
    }


def _next_piano_progressivo(db: Session, tenant_id: int) -> int:
    current = (
        db.query(PianoManutenzione.progressivo)
        .filter(PianoManutenzione.tenant_id == tenant_id)
        .order_by(PianoManutenzione.progressivo.desc())
        .first()
    )
    return (current[0] if current and current[0] else 0) + 1


def _execute_single_import(import_type: str, rows: List[Dict], tenant_id: int, db: Session) -> dict:
    stats: dict[str, int] = {}

    if import_type == "siti":
        stats = {"siti_creati": 0, "siti_esistenti": 0}
        for row in rows:
            existing = _find_sito(db, tenant_id, row.get("nome"))
            if existing:
                stats["siti_esistenti"] += 1
                continue
            db.add(Sito(
                nome=row["nome"].strip(),
                descrizione=row.get("descrizione"),
                ubicazione=row.get("ubicazione"),
                citta=row.get("citta"),
                paese=row.get("paese") or "Italia",
                responsabile=row.get("responsabile"),
                telefono_responsabile=row.get("telefono_responsabile"),
                email_responsabile=row.get("email_responsabile"),
                note=row.get("note"),
                tenant_id=tenant_id,
            ))
            stats["siti_creati"] += 1

    elif import_type == "impianti":
        stats = {"impianti_creati": 0, "impianti_esistenti": 0}
        for row in rows:
            existing = _find_impianto(db, tenant_id, row.get("nome"))
            if existing:
                stats["impianti_esistenti"] += 1
                continue
            sito = _find_sito(db, tenant_id, row.get("sito_nome"))
            db.add(Impianto(
                nome=row["nome"].strip(),
                descrizione=row.get("descrizione"),
                tipologia=row.get("tipologia"),
                note=row.get("note"),
                latitude=_to_float(row.get("latitude")),
                longitude=_to_float(row.get("longitude")),
                sito_id=sito.id if sito else None,
                tenant_id=tenant_id,
            ))
            stats["impianti_creati"] += 1

    elif import_type == "asset":
        stats = {"asset_creati": 0}
        for row in rows:
            impianto = _find_impianto(db, tenant_id, row.get("impianto_nome"))
            db.add(Asset(
                nome=row["nome"].strip(),
                area=row["area"].strip(),
                impianto_id=impianto.id if impianto else None,
                codice=row.get("codice"),
                descrizione=row.get("descrizione"),
                note=row.get("note") or "",
                criticita=_normalize_criticita(row.get("criticita")),
                stato=_normalize_stato(row.get("stato")),
                marca=row.get("marca"),
                modello=row.get("modello"),
                matricola=row.get("matricola"),
                numero_serie=row.get("numero_serie"),
                anno_installazione=_to_int(row.get("anno_installazione")),
                anno_produzione=_to_int(row.get("anno_produzione")),
                fornitore=row.get("fornitore"),
                data_acquisto=_to_date(row.get("data_acquisto")),
                data_scadenza_garanzia=_to_date(row.get("data_scadenza_garanzia")),
                posizione_fisica=row.get("posizione_fisica"),
                limitazioni=row.get("limitazioni"),
                weather_constraint=_normalize_weather(row.get("weather_constraint")),
                fermo_on_schedule=_to_bool(row.get("fermo_on_schedule")),
                note_tecniche=row.get("note_tecniche"),
                vincoli_operativi=row.get("vincoli_operativi"),
                vincoli_manutenzione=row.get("vincoli_manutenzione"),
                tenant_id=tenant_id,
            ))
            stats["asset_creati"] += 1

    elif import_type == "tecnici":
        stats = {"tecnici_creati": 0, "tecnici_esistenti": 0}
        for row in rows:
            if _find_tecnico(db, tenant_id, row):
                stats["tecnici_esistenti"] += 1
                continue
            db.add(Tecnico(
                nome=row["nome"].strip(),
                cognome=row.get("cognome"),
                competenze=row.get("skill"),
                ore_giornaliere=_to_int(row.get("ore_giornaliere")) or 8,
                stato=(row.get("stato") or "in servizio").strip().lower(),
                orario_inizio=row.get("orario_inizio") or "08:00",
                orario_fine=row.get("orario_fine") or "17:00",
                telefono=row.get("telefono"),
                sede_indirizzo=row.get("sede_indirizzo"),
                limitazioni_orarie=row.get("limitazioni_orarie"),
                tenant_id=tenant_id,
            ))
            stats["tecnici_creati"] += 1

    elif import_type == "piani_manutenzione":
        stats = {"piani_creati": 0, "piani_esistenti": 0, "task_creati": 0}
        piani_cache: dict[str, PianoManutenzione] = {}
        for row in rows:
            asset = _find_asset(db, tenant_id, row)
            piano_nome = row["piano_nome"].strip()
            piano = piani_cache.get(piano_nome.lower())
            if not piano:
                piano = (
                    db.query(PianoManutenzione)
                    .filter(PianoManutenzione.tenant_id == tenant_id, PianoManutenzione.nome_codificato.ilike(piano_nome))
                    .first()
                )
            if not piano:
                progressivo = _next_piano_progressivo(db, tenant_id)
                piano = PianoManutenzione(
                    tenant_id=tenant_id,
                    nome_codificato=piano_nome,
                    progressivo=progressivo,
                    descrizione=row.get("piano_descrizione"),
                    asset_id=asset.id if asset else None,
                    stato="attivo",
                )
                db.add(piano)
                db.flush()
                stats["piani_creati"] += 1
            else:
                stats["piani_esistenti"] += 1
            piani_cache[piano_nome.lower()] = piano
            if asset and asset not in piano.assets:
                piano.assets.append(asset)
            db.add(AttivitaManutenzione(
                asset_id=asset.id if asset else None,
                piano_id=piano.id,
                nome=row.get("task_nome"),
                descrizione=row.get("task_descrizione"),
                frequenza_giorni=_to_int(row.get("frequenza_giorni")),
                durata_ore=_to_float(row.get("durata_ore")) or 1,
                priorita=row.get("priorita") or "Media",
                origine="bulk_import",
                generation_mode=row.get("generation_mode") or "manual",
                generate_days_before_due=_to_int(row.get("generate_days_before_due")) or 7,
                source_type="manual_task",
                task_stato="active",
                tenant_id=tenant_id,
            ))
            stats["task_creati"] += 1

    elif import_type == "corsi_attestati":
        stats = {"attestati_creati": 0}
        for row in rows:
            tecnico = _find_tecnico(db, tenant_id, row)
            db.add(Attestato(
                tecnico_id=tecnico.id,
                tenant_id=tenant_id,
                tipo_corso=row["tipo_corso"].strip(),
                ente_certificatore=row.get("ente_certificatore"),
                data_conseguimento=_to_date(row.get("data_conseguimento")),
                data_scadenza=_to_date(row.get("data_scadenza")),
                note=row.get("note"),
            ))
            stats["attestati_creati"] += 1

    return stats


# ── Endpoint: execute ─────────────────────────────────────────────────────────

@router.post("/execute")
async def execute_import(
    file: UploadFile = File(...),
    tenant_id: int = Form(...),
    import_type: str = Form("gerarchia"),
    _sa: dict = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    """
    Esegue l'import massivo dal file Excel nel tenant specificato.

    Logica:
    - SITI: cerca per nome (case-insensitive) nel tenant; se esiste lo usa, altrimenti lo crea.
    - IMPIANTI: cerca per nome nel tenant; se esiste lo usa, altrimenti lo crea collegato al sito.
    - ASSET: crea sempre (non fa dedup su nome perché lo stesso asset può avere più istanze).

    Tutti i record vengono creati con tenant_id = tenant_id specificato.
    L'operazione è atomica (P2-12): un unico commit finale; qualsiasi errore causa rollback completo.
    """
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Il file deve essere in formato .xlsx")

    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(413, f"File troppo grande: massimo {MAX_IMPORT_BYTES // (1024 * 1024)} MB.")

    # Validazione magic bytes: xlsx è famiglia ZIP (PK\x03\x04)
    try:
        validate_upload(content, file.filename, {"xlsx"})
    except ValueError as exc:
        raise HTTPException(400, f"File Excel non valido: {exc}")

    normalized_type = import_type.strip().lower()
    if normalized_type not in HIERARCHY_TYPES:
        rows, row_errors = _read_single_rows(content, normalized_type)
        all_errors = row_errors + _validate_single_rows(normalized_type, rows, db, tenant_id)
        if all_errors:
            raise HTTPException(422, {"errors": all_errors, "message": "Correggere gli errori prima di eseguire l'import"})
        try:
            stats = _execute_single_import(normalized_type, rows, tenant_id, db)
            db.commit()
        except Exception as exc:
            db.rollback()
            logger.error("BulkImport ROLLBACK: type=%s tenant=%d errore=%s", normalized_type, tenant_id, exc)
            raise HTTPException(500, f"Import fallito, nessun dato salvato: {exc}")
        return {
            "success": True,
            "message": f"Import {normalized_type} completato.",
            "stats": stats,
        }

    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Impossibile leggere il file Excel: {e}")

    ws_siti     = wb["SITI"]     if "SITI"     in wb.sheetnames else None
    ws_impianti = wb["IMPIANTI"] if "IMPIANTI" in wb.sheetnames else None
    ws_asset    = wb["ASSET"]    if "ASSET"    in wb.sheetnames else None

    if not ws_siti or not ws_impianti or not ws_asset:
        raise HTTPException(400, "Il file deve contenere i fogli: SITI, IMPIANTI, ASSET")

    siti_rows, siti_errs = _parse_sheet(ws_siti, SITI_COLS)
    impianti_rows, imp_errs = _parse_sheet(ws_impianti, IMPIANTI_COLS)
    asset_rows, asset_errs = _parse_sheet(ws_asset, ASSET_COLS)

    all_errors = siti_errs + imp_errs + asset_errs
    if all_errors:
        raise HTTPException(422, {"errors": all_errors, "message": "Correggere gli errori prima di eseguire l'import"})

    created_siti: Dict[str, int] = {}      # nome_lower → id
    created_impianti: Dict[str, int] = {}  # nome_lower → id

    stats = {"siti_creati": 0, "siti_esistenti": 0,
             "impianti_creati": 0, "impianti_esistenti": 0,
             "asset_creati": 0}

    try:
        # ── SITI ─────────────────────────────────────────────────────────────────
        for r in siti_rows:
            nome_key = r["nome"].strip().lower()
            existing = (
                db.query(Sito)
                .filter(Sito.tenant_id == tenant_id, Sito.nome.ilike(r["nome"].strip()))
                .first()
            )
            if existing:
                created_siti[nome_key] = existing.id
                stats["siti_esistenti"] += 1
            else:
                sito = Sito(
                    nome=r["nome"].strip(),
                    descrizione=r.get("descrizione"),
                    ubicazione=r.get("ubicazione"),
                    citta=r.get("citta"),
                    paese=r.get("paese") or "Italia",
                    responsabile=r.get("responsabile"),
                    telefono_responsabile=r.get("telefono_responsabile"),
                    email_responsabile=r.get("email_responsabile"),
                    note=r.get("note"),
                    tenant_id=tenant_id,
                )
                db.add(sito)
                db.flush()  # ottieni l'id senza commit — visibile nella sessione corrente
                created_siti[nome_key] = sito.id
                stats["siti_creati"] += 1

        # Recupera anche siti già esistenti nel DB (non nel file)
        for s in db.query(Sito).filter(Sito.tenant_id == tenant_id).all():
            created_siti.setdefault(s.nome.strip().lower(), s.id)

        # ── IMPIANTI ──────────────────────────────────────────────────────────────
        for r in impianti_rows:
            nome_key = r["nome"].strip().lower()
            sito_nome_key = (r.get("sito_nome") or "").strip().lower()
            sito_id = created_siti.get(sito_nome_key)

            existing = (
                db.query(Impianto)
                .filter(Impianto.tenant_id == tenant_id, Impianto.nome.ilike(r["nome"].strip()))
                .first()
            )
            if existing:
                created_impianti[nome_key] = existing.id
                stats["impianti_esistenti"] += 1
            else:
                lat = _to_float(r.get("latitude"))
                lon = _to_float(r.get("longitude"))
                impianto = Impianto(
                    nome=r["nome"].strip(),
                    descrizione=r.get("descrizione"),
                    tipologia=r.get("tipologia"),
                    note=r.get("note"),
                    latitude=lat,
                    longitude=lon,
                    sito_id=sito_id,
                    tenant_id=tenant_id,
                )
                db.add(impianto)
                db.flush()
                created_impianti[nome_key] = impianto.id
                stats["impianti_creati"] += 1

        # Recupera anche impianti già esistenti nel DB
        for i in db.query(Impianto).filter(Impianto.tenant_id == tenant_id).all():
            created_impianti.setdefault(i.nome.strip().lower(), i.id)

        # ── ASSET ─────────────────────────────────────────────────────────────────
        for r in asset_rows:
            imp_nome_key = (r.get("impianto_nome") or "").strip().lower()
            impianto_id = created_impianti.get(imp_nome_key)

            asset = Asset(
                nome=r["nome"].strip(),
                area=r["area"].strip(),
                impianto_id=impianto_id,
                codice=r.get("codice"),
                descrizione=r.get("descrizione"),
                note=r.get("note") or "",
                criticita=_normalize_criticita(r.get("criticita")),
                stato=_normalize_stato(r.get("stato")),
                marca=r.get("marca"),
                modello=r.get("modello"),
                matricola=r.get("matricola"),
                numero_serie=r.get("numero_serie"),
                anno_installazione=_to_int(r.get("anno_installazione")),
                anno_produzione=_to_int(r.get("anno_produzione")),
                fornitore=r.get("fornitore"),
                data_acquisto=_to_date(r.get("data_acquisto")),
                data_scadenza_garanzia=_to_date(r.get("data_scadenza_garanzia")),
                posizione_fisica=r.get("posizione_fisica"),
                limitazioni=r.get("limitazioni"),
                weather_constraint=_normalize_weather(r.get("weather_constraint")),
                fermo_on_schedule=_to_bool(r.get("fermo_on_schedule")),
                note_tecniche=r.get("note_tecniche"),
                vincoli_operativi=r.get("vincoli_operativi"),
                vincoli_manutenzione=r.get("vincoli_manutenzione"),
                tenant_id=tenant_id,
            )
            db.add(asset)
            stats["asset_creati"] += 1

        db.commit()  # unico commit atomico — o tutto o niente (P2-12)

    except Exception as exc:
        db.rollback()
        logger.error("BulkImport ROLLBACK: tenant=%d errore=%s", tenant_id, exc)
        raise HTTPException(500, f"Import fallito, nessun dato salvato: {exc}")

    logger.info(
        "BulkImport: tenant=%d — siti +%d (exist %d), impianti +%d (exist %d), asset +%d",
        tenant_id,
        stats["siti_creati"], stats["siti_esistenti"],
        stats["impianti_creati"], stats["impianti_esistenti"],
        stats["asset_creati"],
    )

    return {
        "success": True,
        "message": (
            f"Import completato: {stats['siti_creati']} siti creati, "
            f"{stats['impianti_creati']} impianti creati, "
            f"{stats['asset_creati']} asset creati."
        ),
        "stats": stats,
    }


# ── Helper di conversione ────────────────────────────────────────────────────

def _to_float(val: Optional[str]) -> Optional[float]:
    if not val:
        return None
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _to_int(val: Optional[str]) -> Optional[int]:
    if not val:
        return None
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


def _to_date(val: Optional[str]) -> Optional[date]:
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_bool(val: Optional[str]) -> bool:
    if not val:
        return False
    return str(val).strip().upper() in ("SI", "SÌ", "YES", "1", "TRUE", "VERO")


def _normalize_criticita(val: Optional[str]) -> str:
    v = (val or "media").strip().lower()
    return v if v in ("bassa", "media", "alta") else "media"


def _normalize_stato(val: Optional[str]) -> str:
    v = (val or "service").strip().lower()
    if v in ("operativo", "in servizio", "in_servizio", "service"):
        return "service"
    if v in ("fermo", "fermo prog", "fermo prog.", "fermo programmato", "stopped"):
        return "stopped"
    if v in ("guasto", "fuori servizio", "oos", "out of service", "out_of_service"):
        return "out of service"
    return "service"


WEATHER_VALUES = {"NONE", "NO_RAIN", "NO_WIND", "NO_FROST", "OUTDOOR_ONLY", "INDOOR_ONLY"}

def _normalize_weather(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    v = str(val).strip().upper()
    return v if v in WEATHER_VALUES else None
