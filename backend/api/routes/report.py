"""
Routes: report economico manutenzione (M5.3).
- GET /report/economico        → dati JSON (costo evitato/subito, trend, top asset)
- GET /report/economico/export → file Excel (openpyxl)
"""
import io
import logging
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.core.dependencies import get_db
from backend.core.security import get_current_tenant_id
from backend.db.modelli import Asset, Ticket

logger = logging.getLogger(__name__)

router = APIRouter()


# ── helpers ──────────────────────────────────────────────────────────────────

def _mese_label(dt: datetime) -> str:
    """'2025-06'"""
    return dt.strftime("%Y-%m")


def _mese_display(label: str) -> str:
    """'2025-06' → 'Giu 2025'"""
    mesi = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu",
            "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]
    try:
        y, m = label.split("-")
        return f"{mesi[int(m) - 1]} {y}"
    except Exception:
        return label


def _calcola_costo(ticket: Ticket, asset_map: dict) -> float:
    """Calcola costo fermo per un ticket: durata_ore × costo_orario_fermo asset."""
    asset = asset_map.get(ticket.asset_id)
    if not asset:
        return 0.0
    cof = asset.costo_orario_fermo or 0.0
    ore = ticket.durata_stimata_ore or 0.0
    return ore * cof


# ── endpoint JSON ─────────────────────────────────────────────────────────────

@router.get("/report/economico")
def get_report_economico(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
    mesi: int = 12,
):
    """
    Ritorna il report economico di manutenzione per il tenant:
    - Costo fermo evitato (PM/CM chiusi × costo_orario_fermo asset)
    - Costo fermo subito   (BD chiusi × costo_orario_fermo asset)
    - Trend mensile (ultimi `mesi` mesi, default 12)
    - Top 5 asset per costo subito
    """
    now = datetime.now(tz=timezone.utc)
    da = now - timedelta(days=30 * mesi)

    # Tutti i ticket chiusi nel periodo
    tickets = (
        db.query(Ticket)
        .filter(
            Ticket.tenant_id == tenant_id,
            Ticket.stato == "Chiuso",
            Ticket.deleted_at.is_(None),
            Ticket.created_at >= da,
        )
        .all()
    )

    # Asset map per il tenant (pre-carica tutti per evitare N+1)
    asset_ids = {t.asset_id for t in tickets if t.asset_id}
    assets = db.query(Asset).filter(Asset.id.in_(asset_ids)).all() if asset_ids else []
    asset_map = {a.id: a for a in assets}

    # ── Aggregati totali ──────────────────────────────────────────────────────
    costo_evitato = 0.0   # PM + CM chiusi
    costo_subito   = 0.0  # BD chiusi
    ticket_pm_cm   = 0
    ticket_bd      = 0

    # ── Trend mensile ─────────────────────────────────────────────────────────
    mesi_evitato: dict[str, float] = defaultdict(float)
    mesi_subito:  dict[str, float] = defaultdict(float)

    # ── Top asset ────────────────────────────────────────────────────────────
    asset_subito:  dict[int, float] = defaultdict(float)
    asset_bd_count: dict[int, int]  = defaultdict(int)

    for t in tickets:
        costo = _calcola_costo(t, asset_map)
        mese  = _mese_label(t.created_at) if t.created_at else "?"

        if t.tipo == "BD":
            costo_subito += costo
            ticket_bd    += 1
            mesi_subito[mese]  += costo
            if t.asset_id:
                asset_subito[t.asset_id]   += costo
                asset_bd_count[t.asset_id] += 1
        else:  # PM, CM
            costo_evitato += costo
            ticket_pm_cm  += 1
            mesi_evitato[mese] += costo

    totale = costo_evitato + costo_subito
    ratio_prevenzione = round((costo_evitato / totale * 100) if totale > 0 else 0, 1)

    # ── Costruisci serie trend mensile (12 mesi, anche quelli vuoti) ─────────
    trend: List[dict] = []
    for i in range(mesi - 1, -1, -1):
        dt = now - timedelta(days=30 * i)
        lbl = _mese_label(dt)
        trend.append({
            "mese":     lbl,
            "label":    _mese_display(lbl),
            "evitato":  round(mesi_evitato.get(lbl, 0.0), 2),
            "subito":   round(mesi_subito.get(lbl, 0.0), 2),
        })

    # ── Top 5 asset per costo subito ─────────────────────────────────────────
    top_asset_ids = sorted(asset_subito, key=lambda k: asset_subito[k], reverse=True)[:5]
    top_asset = []
    for aid in top_asset_ids:
        a = asset_map.get(aid)
        top_asset.append({
            "asset_id":    aid,
            "nome":        a.nome if a else f"Asset #{aid}",
            "costo_subito": round(asset_subito[aid], 2),
            "ticket_bd":   asset_bd_count[aid],
            "costo_orario_fermo": (a.costo_orario_fermo or 0.0) if a else 0.0,
        })

    # ── Asset senza costo configurato (warning) ───────────────────────────────
    asset_senza_costo = sum(
        1 for t in tickets
        if t.asset_id and (asset_map.get(t.asset_id) is None
                           or not asset_map[t.asset_id].costo_orario_fermo)
    )

    return {
        "periodo_mesi":        mesi,
        "da":                  da.date().isoformat(),
        "a":                   now.date().isoformat(),
        "costo_fermo_evitato": round(costo_evitato, 2),
        "costo_fermo_subito":  round(costo_subito, 2),
        "totale":              round(totale, 2),
        "ratio_prevenzione":   ratio_prevenzione,
        "ticket_pm_cm_chiusi": ticket_pm_cm,
        "ticket_bd_chiusi":    ticket_bd,
        "trend_mensile":       trend,
        "top_asset":           top_asset,
        "warning_asset_senza_costo": asset_senza_costo,
    }


# ── endpoint Excel export ─────────────────────────────────────────────────────

@router.get("/report/economico/export")
def export_report_excel(
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_current_tenant_id),
):
    """Genera file Excel con il report economico (openpyxl)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=501, detail="openpyxl non disponibile")

    # Ricalcola dati (12 mesi)
    data = get_report_economico(db=db, tenant_id=tenant_id, mesi=12)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Economico"

    # ── Stili ─────────────────────────────────────────────────────────────────
    VERDE  = PatternFill("solid", fgColor="1a3a2a")
    ROSSO  = PatternFill("solid", fgColor="3a1a1a")
    BLU    = PatternFill("solid", fgColor="1a2a3a")
    HEADER = PatternFill("solid", fgColor="1f2937")
    bold   = Font(bold=True, color="FFFFFF", size=11)
    normal = Font(color="FFFFFF", size=10)

    def _h(ws, row, col, val, fill=HEADER, font=bold):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal="center")
        return c

    def _v(ws, row, col, val, fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = normal
        if fmt:
            c.number_format = fmt
        return c

    # ── Intestazione ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"MaintAI — Report Economico Manutenzione ({data['da']} → {data['a']})"
    t.font = Font(bold=True, color="FFFFFF", size=13)
    t.fill = PatternFill("solid", fgColor="0a0f1e")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── KPI principali ────────────────────────────────────────────────────────
    ws.cell(row=3, column=1, value="KPI").font = Font(bold=True, color="AAAAAA", size=9)

    _h(ws, 4, 1, "Costo Fermo EVITATO (€)", VERDE, bold)
    _v(ws, 4, 2, data["costo_fermo_evitato"], "#,##0.00 €")

    _h(ws, 5, 1, "Costo Fermo SUBITO (€)", ROSSO, bold)
    _v(ws, 5, 2, data["costo_fermo_subito"], "#,##0.00 €")

    _h(ws, 6, 1, "Ratio Prevenzione (%)", BLU, bold)
    _v(ws, 6, 2, data["ratio_prevenzione"], "0.0%")

    _h(ws, 7, 1, "Ticket PM/CM chiusi", HEADER, bold)
    _v(ws, 7, 2, data["ticket_pm_cm_chiusi"])

    _h(ws, 8, 1, "Ticket BD chiusi", HEADER, bold)
    _v(ws, 8, 2, data["ticket_bd_chiusi"])

    # ── Trend mensile ──────────────────────────────────────────────────────────
    ws.cell(row=10, column=1, value="TREND MENSILE").font = Font(bold=True, color="AAAAAA", size=9)
    _h(ws, 11, 1, "Mese"); _h(ws, 11, 2, "Evitato (€)"); _h(ws, 11, 3, "Subito (€)")

    for i, row in enumerate(data["trend_mensile"]):
        r = 12 + i
        ws.cell(row=r, column=1, value=row["label"]).font = normal
        _v(ws, r, 2, row["evitato"], "#,##0.00")
        _v(ws, r, 3, row["subito"],  "#,##0.00")

    # ── Top asset ─────────────────────────────────────────────────────────────
    start_top = 12 + len(data["trend_mensile"]) + 2
    ws.cell(row=start_top, column=1, value="TOP ASSET PER COSTO SUBITO").font = Font(bold=True, color="AAAAAA", size=9)
    _h(ws, start_top + 1, 1, "Asset")
    _h(ws, start_top + 1, 2, "Costo Subito (€)")
    _h(ws, start_top + 1, 3, "Ticket BD")
    _h(ws, start_top + 1, 4, "€/ora fermo")

    for i, a in enumerate(data["top_asset"]):
        r = start_top + 2 + i
        ws.cell(row=r, column=1, value=a["nome"]).font = normal
        _v(ws, r, 2, a["costo_subito"],        "#,##0.00")
        _v(ws, r, 3, a["ticket_bd"])
        _v(ws, r, 4, a["costo_orario_fermo"], "#,##0.00")

    # Colonne larghezza
    for col, w in [(1, 32), (2, 20), (3, 15), (4, 18), (5, 12)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # Sfondo foglio scuro
    ws.sheet_properties.tabColor = "0a0f1e"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_file = f"maintai_report_economico_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_file}"'},
    )
